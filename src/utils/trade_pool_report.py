"""
交易对池 Excel 导出与推送工具。
"""

from __future__ import annotations

import smtplib
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import requests


def export_trade_pool_to_excel(pool_data: Iterable[Dict], output_dir: str, prefix: str = "trade_pool") -> str:
    """
    将交易对池导出为 xlsx 文件，返回文件绝对路径。
    运行时依赖 openpyxl；若未安装会抛出可读错误。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl，无法导出 Excel，请先安装 openpyxl") from exc

    rows: List[Dict] = list(pool_data or [])
    output_path = Path(output_dir).expanduser().resolve()
    output_path.mkdir(parents=True, exist_ok=True)

    filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_path / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TradePool"

    headers = ["时间", "交易对", "方向", "价格", "24h涨幅", "评分", "命中次数", "信号理由", "策略来源"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in rows:
        sheet.append([
            item.get("time", ""),
            item.get("symbol", ""),
            item.get("direction", ""),
            float(item.get("price", 0.0) or 0.0),
            float(item.get("change_24h", 0.0) or 0.0),
            float(item.get("score", 0.0) or 0.0),
            int(item.get("hits", 1) or 1),
            str(item.get("reason", "")),
            str(item.get("strategy", "")),
        ])

    widths = {
        "A": 20,
        "B": 20,
        "C": 10,
        "D": 14,
        "E": 12,
        "F": 10,
        "G": 10,
        "H": 60,
        "I": 26,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    workbook.save(file_path)
    return str(file_path)


def send_telegram_document(token: str, chat_id: str, file_path: str, caption: str = ""):
    """通过 Telegram Bot 发送 Excel 文件。"""
    if not token or not chat_id:
        raise ValueError("Telegram token 或 chat_id 为空")

    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as f:
        response = requests.post(
            url,
            data={"chat_id": chat_id, "caption": caption},
            files={"document": (Path(file_path).name, f)},
            timeout=30,
        )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        raise RuntimeError(f"Telegram 推送失败: {payload}")


def send_server_chan_message(send_key: str, title: str, content: str):
    """通过 Server 酱发送消息到个人微信。"""
    if not send_key:
        raise ValueError("Server 酱 SendKey 为空")

    url = f"https://sctapi.ftqq.com/{send_key}.send"
    response = requests.post(
        url,
        data={"title": title, "desp": content},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    # Server酱新版返回码 0 为成功
    if payload.get("code") != 0 and payload.get("data", {}).get("error") != "SUCCESS":
        if payload.get("errno") != 0: # 兼容旧版或不同版本
             raise RuntimeError(f"Server 酱推送失败: {payload}")


def send_email_with_attachment(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    to_email: str,
    file_path: str,
    subject: str,
    body: str,
    use_tls: bool = True,
):
    """通过 SMTP 发送 Excel 附件。"""
    if not smtp_host or not smtp_port or not smtp_user or not smtp_password or not to_email:
        raise ValueError("SMTP 参数不完整")

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_user
    message["To"] = to_email
    message.set_content(body)

    file_name = Path(file_path).name
    with open(file_path, "rb") as f:
        content = f.read()
    message.add_attachment(
        content,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file_name,
    )

    with smtplib.SMTP(smtp_host, int(smtp_port), timeout=30) as server:
        if use_tls:
            server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(message)
