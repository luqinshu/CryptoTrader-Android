"""
重点监控池 UI 页面 - 管理监控交易对，实时展示监控信号
"""

import json
import time
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


from src.monitor.engine import MonitorWorker, SignalType
from src.qt_compat import QApplication, QCheckBox, QColor, QComboBox, QCompleter, QDialog, QDoubleSpinBox, QFormLayout, QFrame, QGridLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMenu, QMessageBox, QPushButton, QScrollArea, QSlider, QSpinBox, QSplitter, QTableWidget, QTableWidgetItem, QTextEdit, QTimer, QVBoxLayout, QWidget, Qt, Signal


# ── 独立信号日志弹窗 ──
class SignalLogWindow(QDialog):
    """可独立调整大小的实时信号日志窗口"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📡 实时信号日志 — CryptoTrader")
        self.setMinimumSize(600, 400)
        self.resize(800, 550)
        self.setStyleSheet("background-color: #0d1117;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # 标题行
        header = QHBoxLayout()
        title = QLabel("📡 实时信号日志")
        title.setStyleSheet("color: #00ff88; font-size: 14px; font-weight: bold;")
        header.addWidget(title)

        self.win_count_label = QLabel("0 条信号")
        self.win_count_label.setStyleSheet("color: #aaaaaa; font-size: 12px;")
        header.addWidget(self.win_count_label)
        header.addStretch()

        clear_btn = QPushButton("清空")
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #444; color: white; padding: 4px 12px;
                border-radius: 3px; font-size: 11px;
            }
            QPushButton:hover { background-color: #555; }
        """)
        clear_btn.clicked.connect(self.clear_log)
        header.addWidget(clear_btn)
        layout.addLayout(header)

        # 日志文本区
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #0a0a0a; color: #cccccc;
                border: 1px solid #333; border-radius: 4px;
                font-family: 'Menlo', 'Monaco'; font-size: 11px;
                padding: 8px;
            }
        """)
        layout.addWidget(self.log_text)

        # 底部状态栏
        status_layout = QHBoxLayout()
        self.status_label = QLabel("🟢 监控中")
        self.status_label.setStyleSheet("color: #00ff88; font-size: 11px;")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()

        always_top_check = QCheckBox("始终置顶")
        always_top_check.setStyleSheet("color: #888888; font-size: 11px;")
        always_top_check.toggled.connect(self._toggle_always_on_top)
        status_layout.addWidget(always_top_check)
        layout.addLayout(status_layout)

        self._signal_count = 0

    def append_log(self, html: str):
        self.log_text.append(html)
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def increment_count(self):
        self._signal_count += 1
        self.win_count_label.setText(f"{self._signal_count} 条信号")

    def set_status(self, text: str, color: str = "#00ff88"):
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color: {color}; font-size: 11px;")

    def clear_log(self):
        self.log_text.clear()
        self._signal_count = 0
        self.win_count_label.setText("0 条信号")

    def _toggle_always_on_top(self, checked: bool):
        if checked:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowStaysOnTopHint)
        self.show()


# ── 信号弹窗提醒 ──
class SignalAlertDialog(QDialog):
    """信号弹窗提醒 - 非模态，自动30秒消失，右上角堆叠排列"""

    _ICONS = {
        '趋势突破': '🚀', '大幅回调': '📉',
        '企稳突破': '🎯', '放量异动': '📊', '动量背离': '⚡',
    }
    _TYPE_COLORS = {
        '趋势突破': '#ff6b35', '大幅回调': '#ff4444',
        '企稳突破': '#00d4aa', '放量异动': '#ffd166', '动量背离': '#c77dff',
    }
    _DIR_COLORS = {'BUY': '#00ff88', 'SHORT': '#ff4444', 'NEUTRAL': '#ffaa00'}
    _DIR_TEXTS  = {'BUY': '做多 ▲', 'SHORT': '做空 ▼', 'NEUTRAL': '中性 ●'}
    _active: list = []  # class-level list for stacking

    def __init__(self, signal, suggestion, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Tool | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_DeleteOnClose, True)
        self.setFixedWidth(350)

        type_color = self._TYPE_COLORS.get(signal.signal_type, '#ff6b35')
        dir_color  = self._DIR_COLORS.get(signal.direction, '#ffffff')
        icon       = self._ICONS.get(signal.signal_type, '📡')
        dir_text   = self._DIR_TEXTS.get(signal.direction, signal.direction)

        self.setWindowTitle(f"{icon} {signal.signal_type} — {signal.inst_id}")
        self.setStyleSheet(f"""
            QDialog {{
                background-color: #111827;
                border: 2px solid {type_color};
            }}
            QLabel {{ background: transparent; }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(6)

        # ── 标题行 ──
        header = QHBoxLayout()
        type_lbl = QLabel(f"{icon} {signal.signal_type}")
        type_lbl.setStyleSheet(f"color: {type_color}; font-size: 14px; font-weight: bold;")
        header.addWidget(type_lbl)
        header.addStretch()
        layout.addLayout(header)

        # ── 交易对 + 方向 + 评分 ──
        info = QHBoxLayout()
        sym_lbl = QLabel(signal.inst_id)
        sym_lbl.setStyleSheet("color: #ffd166; font-size: 13px; font-weight: bold;")
        info.addWidget(sym_lbl)
        dir_lbl = QLabel(dir_text)
        dir_lbl.setStyleSheet(f"color: {dir_color}; font-size: 12px; font-weight: bold;")
        info.addWidget(dir_lbl)
        info.addStretch()
        score_lbl = QLabel(f"★ {signal.score:.0f}分")
        score_lbl.setStyleSheet("color: #ffaa00; font-size: 12px; font-weight: bold;")
        info.addWidget(score_lbl)
        layout.addLayout(info)

        # ── 价格 + 时间 ──
        pt_lbl = QLabel(f"💰 ${signal.price:.4f}   🕐 {signal.timestamp}")
        pt_lbl.setStyleSheet("color: #666666; font-size: 10px;")
        layout.addWidget(pt_lbl)

        # ── 分隔线 ──
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f"border: 1px solid {type_color};")
        layout.addWidget(sep)

        # ── 信号描述 ──
        msg_lbl = QLabel(signal.message)
        msg_lbl.setWordWrap(True)
        msg_lbl.setStyleSheet("color: #cccccc; font-size: 11px;")
        layout.addWidget(msg_lbl)

        # ── 交易建议 ──
        if suggestion:
            sugg_row = QHBoxLayout()
            act_lbl = QLabel(f"⚡ {suggestion['action']}")
            act_lbl.setStyleSheet(
                f"color: {suggestion['action_color']}; font-size: 12px; font-weight: bold;")
            sugg_row.addWidget(act_lbl)
            bias_lbl = QLabel(suggestion['trend_phase'])
            bias_lbl.setStyleSheet(f"color: {suggestion['bias_color']}; font-size: 11px;")
            sugg_row.addWidget(bias_lbl)
            sugg_row.addStretch()
            layout.addLayout(sugg_row)
            if suggestion.get('summary'):
                sum_lbl = QLabel(suggestion['summary'])
                sum_lbl.setWordWrap(True)
                sum_lbl.setStyleSheet("color: #888888; font-size: 10px;")
                layout.addWidget(sum_lbl)

        # ── 关闭按钮 ──
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2a2a2a; color: #aaaaaa;
                padding: 4px 20px; border-radius: 3px; font-size: 11px;
                border: 1px solid #444;
            }
            QPushButton:hover { background-color: #3a3a3a; color: #ffffff; }
        """)
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn, 0, Qt.AlignRight)

        # ── 自动关闭 30 秒 ──
        self._timer = QTimer(self)
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.close)
        self._timer.start(30000)

        # ── 定位：右上角堆叠 ──
        SignalAlertDialog._active.append(self)
        self._reposition()
        self.finished.connect(self._on_closed)

    def _reposition(self):
        screen = QApplication.primaryScreen()
        if not screen:
            return
        geom = screen.availableGeometry()
        self.adjustSize()
        x = geom.right() - self.width() - 20
        idx = SignalAlertDialog._active.index(self)
        y = geom.top() + 40 + idx * (self.height() + 8)
        if y + self.height() > geom.bottom() - 20:
            y = geom.bottom() - self.height() - 20
        self.move(x, y)

    def _on_closed(self):
        try:
            SignalAlertDialog._active.remove(self)
        except ValueError:
            pass


# ── 多周期走势分析器 ──
class MultiTimeframeAnalyzer:
    """
    多时间框走势预判。
    优先使用引擎推送的实时指标（MACD / 布林带 / D1趋势 / RSI），
    作为信号类型的补充而非替代，使预判更贴近真实市场状态。
    """

    @staticmethod
    def analyze(signal, live_data: dict = None) -> dict:
        """
        返回 {'1D', '1H', '3m'} 各周期走势预测。
        live_data: pair_data_updated 推送的实时指标字典（MACD/BB/d1_trend/rsi）。
        """
        sig_type  = signal.signal_type
        direction = signal.direction
        score     = signal.score
        details   = signal.details or {}
        ld = live_data or {}

        d_trend = MultiTimeframeAnalyzer._daily_trend(sig_type, direction, score, ld)
        h_trend = MultiTimeframeAnalyzer._hourly_trend(sig_type, direction, score, details, ld)
        m_trend = MultiTimeframeAnalyzer._minute_trend(sig_type, direction, score, details, ld)
        return {"1D": d_trend, "1H": h_trend, "3m": m_trend}

    @staticmethod
    def from_live_data(live_data: dict) -> dict:
        """
        无信号时，纯粹从实时指标推断走势（定时刷新用途）。
        """
        ld = live_data or {}
        d_trend = MultiTimeframeAnalyzer._daily_trend('', 'NEUTRAL', 0, ld)
        h_trend = MultiTimeframeAnalyzer._hourly_from_indicators(ld)
        m_trend = MultiTimeframeAnalyzer._minute_from_indicators(ld)
        return {"1D": d_trend, "1H": h_trend, "3m": m_trend}

    # ── 日线（D1）：使用引擎计算的 EMA20/EMA50 日线趋势 ─────────────────────
    @staticmethod
    def _daily_trend(sig_type: str, direction: str, score: float, ld: dict) -> tuple:
        d1 = ld.get('d1_trend', '')   # 'bull' / 'bear' / 'sideways' / 'unknown'
        rsi = ld.get('rsi', 50.0)

        if d1 == 'bull':
            if direction in ('BUY', 'NEUTRAL', ''):
                return ("日线多头", "#00ff88")
            else:
                return ("逆势做空⚠", "#ffaa00")   # 信号方向与日线相反，风险提示
        if d1 == 'bear':
            if direction in ('SHORT', 'NEUTRAL', ''):
                return ("日线空头", "#ff4444")
            else:
                return ("逆势做多⚠", "#ffaa00")
        if d1 == 'sideways':
            return ("日线横盘", "#aaaaaa")

        # D1数据不足时降级到原有逻辑
        if direction == "BUY"   and score >= 75:  return ("看多", "#00ff88")
        if direction == "BUY":                    return ("偏多", "#8fd3ff")
        if direction == "SHORT" and score >= 75:  return ("看空", "#ff4444")
        if direction == "SHORT":                  return ("偏空", "#ffaa00")
        return ("震荡", "#aaaaaa")

    # ── 小时线（H1）：使用 MACD 状态 + 信号类型融合 ─────────────────────────
    @staticmethod
    def _hourly_trend(sig_type: str, direction: str, score: float,
                      details: dict, ld: dict) -> tuple:
        macd  = ld.get('macd', {})
        cross = macd.get('cross', 'none')
        hist  = macd.get('hist', 0.0)
        growing = macd.get('hist_growing', False)
        above_zero = macd.get('above_zero', False)

        # ① MACD 金叉/死叉是最强即时信号
        if cross == 'golden':
            return ("MACD金叉🔰", "#00ff88")
        if cross == 'death':
            return ("MACD死叉☠", "#ff4444")

        # ② 结合信号类型
        if sig_type == "趋势突破":
            extra = "  MACD助力" if hist > 0 and growing else ""
            return (f"突破{extra}", "#00ff88") if direction == "BUY" else (f"破位{extra}", "#ff4444")
        if sig_type == "大幅回调":
            return ("牛市回调", "#8fd3ff") if direction == "BUY" else ("熊市反弹", "#ffaa00")
        if sig_type == "企稳突破":
            return ("企稳上破", "#00ff88") if direction == "BUY" else ("企稳下破", "#ff4444")
        if sig_type == "放量异动":
            vr = details.get("volume_ratio", 1.0)
            return (("放量拉升", "#00ff88") if direction == "BUY" else ("放量砸盘", "#ff4444")) \
                if vr >= 3.0 else ("放量波动", "#ffaa00")
        if sig_type == "动量背离":
            return ("顶背离⚠", "#ff6666") if direction == "SHORT" else ("底背离✅", "#00ff88")

        # ③ 无信号时用 MACD 方向
        return MultiTimeframeAnalyzer._hourly_from_indicators(ld)

    @staticmethod
    def _hourly_from_indicators(ld: dict) -> tuple:
        """纯指标推断小时线走势"""
        macd = ld.get('macd', {})
        hist = macd.get('hist', 0.0)
        growing = macd.get('hist_growing', False)
        above_zero = macd.get('above_zero', False)
        rsi = ld.get('rsi', 50.0)
        if hist > 0 and growing and above_zero:
            return ("多头动量↑", "#00ff88")
        if hist > 0 and above_zero:
            return ("偏多", "#8fd3ff")
        if hist < 0 and growing and not above_zero:
            return ("空头动量↓", "#ff4444")
        if hist < 0 and not above_zero:
            return ("偏空", "#ffaa00")
        if rsi >= 70:
            return ("超买区间", "#ffaa00")
        if rsi <= 30:
            return ("超卖区间", "#8fd3ff")
        return ("震荡整理", "#aaaaaa")

    # ── 3分钟（M3）：使用布林带位置 + 量价短期动态 ──────────────────────────
    @staticmethod
    def _minute_trend(sig_type: str, direction: str, score: float,
                      details: dict, ld: dict) -> tuple:
        bb = ld.get('bb', {})
        squeeze   = bb.get('squeeze', False)
        above_up  = bb.get('above_upper', False)
        below_lo  = bb.get('below_lower', False)
        bb_pos    = bb.get('position', 0.5)
        bw        = bb.get('bandwidth', 0.0)

        # ① BB突破上轨/下轨
        if above_up:
            return ("突破上轨", "#00ff88") if direction != 'SHORT' else ("超买风险⚠", "#ffaa00")
        if below_lo:
            return ("跌破下轨", "#ff4444") if direction != 'BUY' else ("超卖机会", "#8fd3ff")

        # ② BB收窄蓄势（breakout 前兆）
        if squeeze:
            return ("带宽收窄蓄力⚡", "#ffff66")

        # ③ 信号类型优先
        if sig_type == "放量异动":
            change = details.get("price_change_pct", 0)
            vr     = details.get("volume_ratio", 1.0)
            if abs(change) >= 3 and vr >= 2.5:
                return ("急涨", "#00ff88") if change > 0 else ("急跌", "#ff4444")
            return ("短线异动", "#ffaa00")
        if sig_type == "趋势突破":
            return ("快涨", "#00ff88") if direction == "BUY" else ("快跌", "#ff4444")
        if sig_type == "企稳突破":
            return ("启动", "#8fd3ff")
        if sig_type == "动量背离":
            return ("动量衰竭", "#ffaa00")

        return MultiTimeframeAnalyzer._minute_from_indicators(ld)

    @staticmethod
    def _minute_from_indicators(ld: dict) -> tuple:
        """纯指标推断3分钟走势"""
        bb = ld.get('bb', {})
        pos = bb.get('position', 0.5)
        squeeze = bb.get('squeeze', False)
        if squeeze:
            return ("蓄势待发⚡", "#ffff66")
        if pos >= 0.85:
            return ("近上轨偏强", "#8fd3ff")
        if pos <= 0.15:
            return ("近下轨偏弱", "#ffaa00")
        vol_ratio = ld.get('volume_ratio', 1.0)
        if vol_ratio >= 2.5:
            return ("短线放量", "#ffaa00")
        return ("盘整", "#888888")


# ── 交易建议引擎 ──
class TradingSuggester:
    """根据监测信号生成交易建议"""

    @staticmethod
    def generate(signal) -> dict:
        """分析信号并返回建议"""
        sig_type = signal.signal_type
        direction = signal.direction
        score = signal.score
        details = signal.details or {}

        result = {
            "action": "观望",
            "action_color": "#aaaaaa",
            "trend_phase": "震荡",
            "bias": "中性",
            "bias_color": "#aaaaaa",
            "summary": "",
        }

        if sig_type == SignalType.TREND_BREAKOUT:
            TradingSuggester._suggest_breakout(signal, result, details)
        elif sig_type == SignalType.DEEP_PULLBACK:
            TradingSuggester._suggest_pullback(signal, result, details)
        elif sig_type == SignalType.STABILIZATION_BREAKOUT:
            TradingSuggester._suggest_stabilization(signal, result, details)
        elif sig_type == SignalType.VOLUME_SURGE:
            TradingSuggester._suggest_volume_surge(signal, result, details)
        elif sig_type == SignalType.MOMENTUM_DIVERGENCE:
            TradingSuggester._suggest_divergence(signal, result, details)

        return result

    @staticmethod
    def _suggest_breakout(signal, result, details):
        vol_ratio = details.get("volume_ratio", 1.0)
        result["trend_phase"] = "突破"

        if signal.direction == "BUY":
            result["bias"] = "看多"
            result["bias_color"] = "#00ff88"
            if signal.score >= 80:
                result["action"] = "加仓"
                result["action_color"] = "#00ff88"
                result["summary"] = f"放量{vol_ratio:.1f}x突破阻力 · 强势信号 · 建议顺势加仓"
            elif signal.score >= 65:
                result["action"] = "关注做多"
                result["action_color"] = "#8fd3ff"
                result["summary"] = f"突破阻力但量能{vol_ratio:.1f}x一般 · 关注回踩确认后入场"
            else:
                result["action"] = "观望偏多"
                result["action_color"] = "#cccccc"
                result["summary"] = f"突破信号偏弱 · 等待进一步确认"
        else:
            result["bias"] = "看空"
            result["bias_color"] = "#ff6666"
            if signal.score >= 80:
                result["action"] = "清仓/做空"
                result["action_color"] = "#ff4444"
                result["summary"] = f"放量{vol_ratio:.1f}x跌破支撑 · 强势空头 · 建议清仓或做空"
            elif signal.score >= 65:
                result["action"] = "减仓"
                result["action_color"] = "#ffaa00"
                result["summary"] = f"跌破支撑量能{vol_ratio:.1f}x · 建议部分减仓控制风险"
            else:
                result["action"] = "观望偏空"
                result["action_color"] = "#cccccc"
                result["summary"] = "破位信号偏弱 · 观察是否有效跌破"

    @staticmethod
    def _suggest_pullback(signal, result, details):
        drawdown = details.get("drawdown_pct", 0)
        rsi = details.get("rsi", 50)

        if signal.direction == "BUY":
            result["trend_phase"] = "牛市回调"
            result["bias"] = "看多"
            result["bias_color"] = "#00ff88"
            if signal.score >= 75:
                result["action"] = "加仓"
                result["action_color"] = "#00ff88"
                result["summary"] = f"回调{drawdown:.1f}%至RSI={rsi:.0f}超卖区 · 主趋势向上 · 优质加仓点"
            else:
                result["action"] = "关注加仓"
                result["action_color"] = "#8fd3ff"
                result["summary"] = f"回调{drawdown:.1f}% · 主趋势向上 · 可逢低分批建仓"
        else:
            result["trend_phase"] = "熊市反弹"
            result["bias"] = "看空"
            result["bias_color"] = "#ff6666"
            if signal.score >= 75:
                result["action"] = "减仓"
                result["action_color"] = "#ffaa00"
                result["summary"] = f"反弹{drawdown:.1f}%至RSI={rsi:.0f} · 主趋势向下 · 反弹是减仓机会"
            else:
                result["action"] = "观望偏空"
                result["action_color"] = "#cccccc"
                result["summary"] = f"反弹{drawdown:.1f}% · 主趋势向下 · 不宜追涨"

    @staticmethod
    def _suggest_stabilization(signal, result, details):
        cons_range = details.get("consolidation_range_pct", 0)
        vol_ratio = details.get("volume_ratio", 1.0)

        if signal.direction == "BUY":
            result["trend_phase"] = "企稳突破"
            result["bias"] = "看多"
            result["bias_color"] = "#00ff88"
            if signal.score >= 75:
                result["action"] = "加仓"
                result["action_color"] = "#00ff88"
                result["summary"] = f"横盘{cons_range:.1f}%后放量{vol_ratio:.1f}x突破 · 趋势启动信号 · 建议入场"
            else:
                result["action"] = "关注做多"
                result["action_color"] = "#8fd3ff"
                result["summary"] = f"企稳突破 · 振幅{cons_range:.1f}% · 可轻仓试多"
        else:
            result["trend_phase"] = "企稳破位"
            result["bias"] = "看空"
            result["bias_color"] = "#ff6666"
            if signal.score >= 75:
                result["action"] = "减仓"
                result["action_color"] = "#ffaa00"
                result["summary"] = f"横盘{cons_range:.1f}%后放量{vol_ratio:.1f}x破位 · 趋势转空 · 建议减仓"
            else:
                result["action"] = "关注做空"
                result["action_color"] = "#8fd3ff"
                result["summary"] = f"企稳破位 · 振幅{cons_range:.1f}% · 可轻仓试空"

    @staticmethod
    def _suggest_volume_surge(signal, result, details):
        vol_ratio = details.get("volume_ratio", 1.0)
        change = details.get("price_change_pct", 0)
        result["trend_phase"] = "放量异动"

        if signal.direction == "BUY":
            result["bias"] = "震荡偏多"
            result["bias_color"] = "#8fd3ff"
            result["action"] = "关注"
            result["action_color"] = "#ffaa00"
            result["summary"] = f"放量{vol_ratio:.1f}x涨{change:+.2f}% · 关注能否持续 · 连续放量可追"
        else:
            result["bias"] = "震荡偏空"
            result["bias_color"] = "#ffaa00"
            result["action"] = "警惕"
            result["action_color"] = "#ff6666"
            result["summary"] = f"放量{vol_ratio:.1f}x跌{change:+.2f}% · 注意风险 · 可能加速下跌"

    @staticmethod
    def _suggest_divergence(signal, result, details):
        result["trend_phase"] = "背离反转"

        if signal.direction == "SHORT":
            result["bias"] = "看空"
            result["bias_color"] = "#ff6666"
            result["action"] = "减仓"
            result["action_color"] = "#ffaa00"
            result["summary"] = "顶背离 · 上涨动能衰竭 · 建议减仓锁定利润"
        else:
            result["bias"] = "看多"
            result["bias_color"] = "#00ff88"
            result["action"] = "关注反转"
            result["action_color"] = "#8fd3ff"
            result["summary"] = "底背离 · 下跌动能衰竭 · 关注反转信号确认"


class MonitorPoolPage(QWidget):
    """重点监控池页面"""

    telegram_alert_requested = Signal(str, str)  # (chat_id, message)

    def __init__(self, okx_client=None, trade_executor=None, trade_settings_manager=None):
        super().__init__()
        self.okx_client = okx_client
        self.trade_executor = trade_executor
        self.trade_settings_manager = trade_settings_manager
        self.monitor_worker: MonitorWorker = None
        self._signals_log: List[Dict] = []
        self._max_log_entries = 500
        self._swap_pairs: List[str] = []  # 所有可用的 SWAP 交易对
        self._signal_log_window: SignalLogWindow = None  # 独立弹出窗口
        self._pair_trend_cache: Dict[str, dict] = {}   # inst_id -> {"1D": (trend,color), ...}
        self._pair_live_data:  Dict[str, dict] = {}   # inst_id -> 最新实时指标(MACD/BB/D1/RSI)
        self._pinned_pairs: set = set()  # 置顶交易对
        self._config_dir = Path(__file__).resolve().parent.parent
        self._pool_config_path = self._config_dir / "monitor_pool_config.json"
        self._col_resize_timer = None
        self._auto_trade_cooldowns: Dict[str, float] = {}
        # 扫描驱动自动交易编排器引用（by set_scan_auto_trader()），None=未启用
        self._scan_auto_trader_ref = None

        self.init_ui()
        self.load_config()

        # 异步加载可用交易对列表
        QTimer.singleShot(100, self._load_swap_pairs)

    def init_ui(self):
        # ── 页面整体深色背景 ──
        self.setStyleSheet("background-color: #0d1117;")

        # ── 页面级滚动区域 ──
        page_scroll = QScrollArea()
        page_scroll.setWidgetResizable(True)
        page_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        page_scroll.setFrameShape(QFrame.NoFrame)
        page_scroll.setStyleSheet("""
            QScrollArea { background-color: #0d1117; border: none; }
            QScrollBar:vertical {
                background: #111111; width: 12px; margin: 0; border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #555555; border-radius: 6px; min-height: 50px;
            }
            QScrollBar::handle:vertical:hover { background: #888888; }
            QScrollBar::handle:vertical:pressed { background: #aaaaaa; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal { height: 0; }
        """)

        self._page_wrapper = QWidget()
        self._page_wrapper.setMinimumHeight(850)
        outer_layout = QVBoxLayout(self._page_wrapper)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        page_scroll.setWidget(self._page_wrapper)

        page_root_layout = QVBoxLayout(self)
        page_root_layout.setContentsMargins(0, 0, 0, 0)
        page_root_layout.addWidget(page_scroll)



        # ── 顶部控制栏 ──
        control_group = QGroupBox("重点监控池 - 实时监测")
        control_group.setStyleSheet("""
            QGroupBox {
                color: #888888; font-weight: bold; font-size: 11px;
                border: 1px solid #333; border-radius: 4px;
                padding: 4px; margin-top: 2px;
            }
            QGroupBox::title { padding: 0 6px; }
        """)
        control_layout = QVBoxLayout(control_group)
        control_layout.setSpacing(3)
        control_layout.setContentsMargins(4, 4, 4, 4)

        # 标题行
        title_row = QHBoxLayout()
        title = QLabel("🔥 重点监控池")
        title.setStyleSheet("color: #ff6b35; font-size: 12px; font-weight: bold;")
        title_row.addWidget(title)
        title_row.addStretch()

        # 启动/停止按钮
        self.start_btn = QPushButton("▶ 启动监控")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745; color: white; font-weight: bold;
                padding: 6px 16px; border-radius: 4px;
            }
            QPushButton:hover { background-color: #34d058; }
        """)
        self.start_btn.clicked.connect(self.toggle_monitor)
        title_row.addWidget(self.start_btn)

        self.status_label = QLabel("⏸ 已停止")
        self.status_label.setStyleSheet("color: #aaaaaa; font-size: 12px;")
        title_row.addWidget(self.status_label)
        control_layout.addLayout(title_row)

        # 添加交易对行 —— 可搜索下拉框
        add_row = QHBoxLayout()
        add_row.setSpacing(6)

        self.add_input = QComboBox()
        self.add_input.setEditable(True)
        self.add_input.setInsertPolicy(QComboBox.NoInsert)
        self.add_input.setPlaceholderText("输入交易对名称搜索，如 BTC、ETH...")
        self.add_input.setStyleSheet("""
            QComboBox {
                background-color: #1e1e1e; color: #ffffff;
                border: 1px solid #444; border-radius: 4px; padding: 6px;
            }
            QComboBox:focus { border: 1px solid #ff6b35; }
            QComboBox QAbstractItemView {
                background-color: #1e1e1e; color: #dddddd;
                border: 1px solid #444; selection-background-color: #2b5d87;
                outline: none; max-height: 200px;
            }
            QComboBox::drop-down { border: none; width: 20px; }
            QComboBox::down-arrow { image: none; }
        """)
        # 自动补全器
        self._pair_completer = QCompleter([], self.add_input)
        self._pair_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self._pair_completer.setFilterMode(Qt.MatchContains)
        self.add_input.setCompleter(self._pair_completer)
        self.add_input.lineEdit().returnPressed.connect(self.add_pair_from_input)
        add_row.addWidget(self.add_input, 1)

        add_btn = QPushButton("➕ 添加监控")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff6b35; color: white; font-weight: bold;
                padding: 6px 14px; border-radius: 4px;
            }
            QPushButton:hover { background-color: #ff8555; }
        """)
        add_btn.clicked.connect(self.add_pair_from_input)
        add_row.addWidget(add_btn)

        refresh_btn = QPushButton("🔄 刷新列表")
        refresh_btn.setToolTip("从 OKX 重新加载所有 USDT 永续合约交易对")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3a3a3a; color: #cccccc;
                padding: 6px 12px; border-radius: 4px; border: 1px solid #444;
            }
            QPushButton:hover { background-color: #4a4a4a; color: #ffffff; }
        """)
        refresh_btn.clicked.connect(self._load_swap_pairs)
        add_row.addWidget(refresh_btn)

        # 快捷添加按钮
        quick_add = QHBoxLayout()
        quick_add.setSpacing(4)
        for pair in ['BTC-USDT-SWAP', 'ETH-USDT-SWAP', 'SOL-USDT-SWAP',
                      'DOGE-USDT-SWAP', 'BNB-USDT-SWAP', 'PEPE-USDT-SWAP']:
            btn = QPushButton(pair.replace('-USDT-SWAP', ''))
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #2a2a2a; color: #cccccc;
                    border: 1px solid #444; border-radius: 3px;
                    padding: 3px 10px; font-size: 11px;
                }
                QPushButton:hover {
                    background-color: #3a3a3a; border: 1px solid #ff6b35; color: #ff6b35;
                }
            """)
            btn.clicked.connect(lambda _, p=pair: self._quick_add(p))
            quick_add.addWidget(btn)
        quick_add.addStretch()
        add_row_section = QVBoxLayout()
        add_row_section.addLayout(add_row)
        add_row_section.addLayout(quick_add)
        control_layout.addLayout(add_row_section)

        outer_layout.addWidget(control_group)

        # ── 主分割区域 ──
        splitter = QSplitter(Qt.Vertical)
        splitter.setHandleWidth(10)
        splitter.setChildrenCollapsible(False)
        splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #3a3a3a;
                border-top: 1px solid #555;
                border-bottom: 1px solid #555;
            }
            QSplitter::handle:hover {
                background-color: #ff6b35;
                border-top: 1px solid #ff6b35;
                border-bottom: 1px solid #ff6b35;
            }
            QSplitter::handle:pressed {
                background-color: #ff4500;
            }
        """)

        # 上半部分：监控列表
        list_widget = QWidget()
        list_layout = QVBoxLayout(list_widget)
        list_layout.setContentsMargins(0, 0, 0, 0)

        list_header = QHBoxLayout()
        list_title = QLabel("📋 监控列表")
        list_title.setStyleSheet("color: #ffd166; font-size: 11px; font-weight: bold;")
        list_header.addWidget(list_title)

        self.pair_count_label = QLabel("0 个交易对")
        self.pair_count_label.setStyleSheet("color: #aaaaaa;")
        list_header.addWidget(self.pair_count_label)
        list_header.addStretch()

        remove_selected_btn = QPushButton("🗑 移除选中")
        remove_selected_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545; color: white; font-weight: bold;
                padding: 4px 12px; border-radius: 3px;
            }
            QPushButton:hover { background-color: #e8445a; }
        """)
        remove_selected_btn.clicked.connect(self.remove_selected_pair)
        list_header.addWidget(remove_selected_btn)
        list_layout.addLayout(list_header)

        # 监控列表表格
        self.pair_table = QTableWidget()
        self.pair_table.setColumnCount(8)
        self.pair_table.setHorizontalHeaderLabels([
            "交易对", "最新价", "涨跌24h%", "RSI 1H", "量比", "状态", "最近信号", "走势预判"
        ])
        self.pair_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.pair_table.horizontalHeader().setMinimumSectionSize(50)
        # 设置合理的默认列宽
        col_widths = [120, 90, 80, 60, 50, 70, 100, 160]
        for i, w in enumerate(col_widths):
            self.pair_table.setColumnWidth(i, w)
        self.pair_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.pair_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.pair_table.setStyleSheet("""
            QTableWidget {
                background-color: #1a1a1a; color: #dddddd;
                border: 1px solid #333; border-radius: 4px;
                gridline-color: #2a2a2a;
            }
            QTableWidget::item { padding: 2px 4px; }
            QTableWidget::item:selected {
                background-color: #2b5d87; color: #ffffff;
            }
            QHeaderView::section {
                background-color: #252525; color: #aaaaaa;
                border: 1px solid #333; padding: 2px;
                font-weight: bold; font-size: 10px;
            }
        """)
        self.pair_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.pair_table.customContextMenuRequested.connect(self.show_pair_menu)
        # 列宽变化时自动保存
        self.pair_table.horizontalHeader().sectionResized.connect(self._on_column_resized)
        list_layout.addWidget(self.pair_table)

        splitter.addWidget(list_widget)

        # 下半部分：信号日志
        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)

        log_header = QHBoxLayout()
        log_title = QLabel("📡 实时信号日志")
        log_title.setStyleSheet("color: #00ff88; font-size: 11px; font-weight: bold;")
        log_header.addWidget(log_title)

        self.signal_count_label = QLabel("0 条信号")
        self.signal_count_label.setStyleSheet("color: #aaaaaa;")
        log_header.addWidget(self.signal_count_label)
        log_header.addStretch()

        self.telegram_checkbox = QCheckBox("转发到Telegram")
        self.telegram_checkbox.setStyleSheet("color: #8fd3ff;")
        self.telegram_checkbox.setChecked(True)
        log_header.addWidget(self.telegram_checkbox)

        popup_log_btn = QPushButton("📡 弹出独立窗口")
        popup_log_btn.setToolTip("在可调整大小的独立窗口中查看实时信号日志")
        popup_log_btn.setStyleSheet("""
            QPushButton {
                background-color: #1a3a2a; color: #00ff88; font-weight: bold;
                border: 1px solid #2a5a3a; border-radius: 3px;
                padding: 4px 12px; font-size: 11px;
            }
            QPushButton:hover { background-color: #2a5a3a; color: white; }
        """)
        popup_log_btn.clicked.connect(self._open_signal_log_window)
        log_header.addWidget(popup_log_btn)

        clear_log_btn = QPushButton("清空日志")
        clear_log_btn.setStyleSheet("""
            QPushButton {
                background-color: #444; color: white; padding: 4px 12px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #555; }
        """)
        clear_log_btn.clicked.connect(self.clear_signal_log)
        log_header.addWidget(clear_log_btn)
        log_layout.addLayout(log_header)

        self.signal_log = QTextEdit()
        self.signal_log.setReadOnly(True)
        self.signal_log.setStyleSheet("""
            QTextEdit {
                background-color: #111111; color: #cccccc;
                border: 1px solid #333; border-radius: 4px;
                font-family: 'Menlo', 'Monaco'; font-size: 11px;
                padding: 6px;
            }
        """)
        log_layout.addWidget(self.signal_log)

        splitter.addWidget(log_widget)
        splitter.setStretchFactor(0, 3)  # 监控列表占 3/4
        splitter.setStretchFactor(1, 1)  # 信号日志占 1/4
        splitter.setSizes([720, 280])

        # ── 底部滚动区域（监测参数 + 信号应对）──
        self.bottom_scroll = QScrollArea()
        self.bottom_scroll.setWidgetResizable(True)
        self.bottom_scroll.setFrameShape(QFrame.NoFrame)
        self.bottom_scroll.setMinimumHeight(60)
        self.bottom_scroll.setStyleSheet("""
            QScrollArea { background-color: transparent; border: none; }
            QScrollBar:vertical {
                background: #1a1a1a; width: 6px; margin: 0; border-radius: 3px;
            }
            QScrollBar::handle:vertical {
                background: #555; border-radius: 3px; min-height: 20px;
            }
            QScrollBar::handle:vertical:hover { background: #777; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
        """)
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(8, 0, 8, 0)
        bottom_layout.setSpacing(4)

        # ── 监测参数区 (滑轨版) ──
        settings_group = QGroupBox("⚙️ 监测参数（拖动滑轨调整）")
        settings_group.setStyleSheet("""
            QGroupBox {
                color: #999999; font-weight: bold; font-size: 11px;
                border: 1px solid #333; border-radius: 4px;
                padding: 4px; margin-top: 2px;
            }
            QGroupBox::title { padding: 0 6px; }
            QGroupBox QLabel { color: #aaaaaa; font-size: 11px; }
        """)
        settings_main = QVBoxLayout(settings_group)
        settings_main.setSpacing(2)
        settings_main.setContentsMargins(6, 8, 6, 6)

        self._sliders: Dict[str, dict] = {}

        # ── 两列滑轨网格 ──
        cols_widget = QWidget()
        cols_layout = QHBoxLayout(cols_widget)
        cols_layout.setContentsMargins(0, 0, 0, 0)
        cols_layout.setSpacing(16)

        _slider_style = """
            QSlider { margin: 3px 0; }
            QSlider::groove:horizontal {
                height: 4px; background: #2a2a2a; border-radius: 2px;
            }
            QSlider::handle:horizontal {
                width: 13px; height: 13px; margin: -5px 0;
                background: #ff6b35; border-radius: 7px;
                border: 1px solid #ff8855;
            }
            QSlider::handle:horizontal:hover { background: #ff8a55; }
            QSlider::sub-page:horizontal { background: #ff6b35; border-radius: 2px; }
        """

        def _make_grid():
            g = QGridLayout()
            g.setSpacing(1)
            g.setColumnMinimumWidth(0, 105)
            g.setColumnStretch(1, 1)
            g.setColumnMinimumWidth(2, 62)
            return g

        left_grid  = _make_grid()
        right_grid = _make_grid()

        def _add_row(grid, row, key, name, lo, hi, default, scale, unit, fmt, influence):
            name_lbl = QLabel(name + ':')
            name_lbl.setStyleSheet("color: #cccccc; font-size: 11px;")
            grid.addWidget(name_lbl, row, 0)

            slider = QSlider(Qt.Horizontal)
            slider.setRange(lo, hi)
            slider.setValue(default)
            slider.setStyleSheet(_slider_style)
            grid.addWidget(slider, row, 1)

            def _fmt_val(v, sc=scale, f=fmt):
                rv = v / sc
                return f"{int(rv):{f}}" if f == 'd' else f"{rv:{f}}"

            val_lbl = QLabel(f"{_fmt_val(default)} {unit}")
            val_lbl.setStyleSheet("color: #ffd166; font-size: 10px; min-width: 60px;")
            val_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            grid.addWidget(val_lbl, row, 2)

            def _update(v, lbl=val_lbl, fv=_fmt_val, u=unit):
                lbl.setText(f"{fv(v)} {u}")
            slider.valueChanged.connect(_update)

            inf_lbl = QLabel(influence)
            inf_lbl.setStyleSheet("color: #505050; font-size: 9px; padding-left: 2px;")
            grid.addWidget(inf_lbl, row + 1, 0, 1, 3)

            self._sliders[key] = {
                'slider': slider, 'label': val_lbl,
                'scale': scale, 'fmt': fmt, 'unit': unit,
            }

        # ── 左列 5 个参数 ──
        _add_row(left_grid, 0, 'check_interval_sec', '检查间隔',
                 15, 3600, 60, 1, '秒', 'd',
                 '↓越小=响应越快  ·  [全部信号] 检测频率★★')
        _add_row(left_grid, 2, 'trend_breakout_pct', '突破幅度阈值',
                 5, 100, 25, 10, '%', '.1f',
                 '↑越大=信号越稀少可靠  ·  [趋势突破]★★★')
        _add_row(left_grid, 4, 'trend_breakout_volume_ratio', '突破放量倍数',
                 10, 50, 15, 10, 'x', '.1f',
                 '↑越大=要求放量更多  ·  [趋势突破]★★★  [企稳突破]★★')
        _add_row(left_grid, 6, 'deep_pullback_pct', '深度回调阈值',
                 10, 150, 50, 10, '%', '.1f',
                 '↑越大=回调幅度要求更深  ·  [大幅回调]★★★')
        _add_row(left_grid, 8, 'stabilization_max_range_pct', '企稳最大振幅',
                 3, 30, 12, 10, '%', '.1f',
                 '↓越小=要求横盘越紧密  ·  [企稳突破]★★★')

        # ── 右列 4 个参数 ──
        _add_row(right_grid, 0, 'stabilization_breakout_volume_ratio', '企稳突破放量',
                 10, 50, 18, 10, 'x', '.1f',
                 '↑越大=突破需更大量能  ·  [企稳突破]★★★')
        _add_row(right_grid, 2, 'volume_surge_ratio', '放量异动倍数',
                 15, 100, 30, 10, 'x', '.1f',
                 '↑越大=异动阈值更高  ·  [放量异动]★★★  [趋势突破]★')
        _add_row(right_grid, 4, 'min_signal_score', '最低信号分',
                 30, 95, 60, 1, '分', 'd',
                 '↑越大=过滤越严格  ·  [全部信号]★★★')
        _add_row(right_grid, 6, 'signal_cooldown_min', '信号冷却时间',
                 5, 120, 15, 1, '分钟', 'd',
                 '↑越大=重复信号越少  ·  [全部信号]★★')

        cols_layout.addLayout(left_grid, 1)
        cols_layout.addLayout(right_grid, 1)
        settings_main.addWidget(cols_widget)

        # ── 底部控制行：应用按钮 + 预设 ──
        bottom_ctrl = QHBoxLayout()
        bottom_ctrl.setSpacing(6)

        apply_btn = QPushButton("应用参数")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0077cc; color: white; font-weight: bold;
                padding: 3px 14px; border-radius: 3px; font-size: 11px;
            }
            QPushButton:hover { background-color: #0088ee; }
        """)
        apply_btn.clicked.connect(self.apply_settings)
        bottom_ctrl.addWidget(apply_btn)
        bottom_ctrl.addSpacing(10)

        preset_label = QLabel("风格预设:")
        preset_label.setStyleSheet("color: #8fd3ff; font-size: 11px; font-weight: bold;")
        bottom_ctrl.addWidget(preset_label)

        presets = [
            ("🛡 保守", "conservative", "#2d6a4f",
             "高阈值 · 低频率 · 适合 BTC/ETH 等主流币"),
            ("⚖ 中性", "neutral", "#3a6b8f",
             "均衡配置 · 适中频率 · 通用推荐"),
            ("⚡ 激进", "aggressive", "#b8860b",
             "低阈值 · 高频率 · 适合波动大的山寨币"),
        ]
        self._preset_buttons = {}
        for label, key, color, tip in presets:
            btn = QPushButton(label)
            btn.setToolTip(tip)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color}; color: white; font-weight: bold;
                    padding: 2px 8px; border-radius: 3px; font-size: 10px;
                    border: 1px solid rgba(255,255,255,0.15);
                }}
                QPushButton:hover {{ border: 1px solid rgba(255,255,255,0.4); }}
            """)
            btn.clicked.connect(lambda _, k=key: self._apply_preset(k))
            self._preset_buttons[key] = btn
            bottom_ctrl.addWidget(btn)

        self._active_preset = "neutral"
        bottom_ctrl.addStretch()
        settings_main.addLayout(bottom_ctrl)

        bottom_layout.addWidget(settings_group)

        # ── 信号应对措施区 ──
        action_group = QGroupBox("🎯 信号应对措施 (选择触发信号时的操作)")
        action_group.setStyleSheet("""
            QGroupBox {
                color: #ffaa00; font-weight: bold; font-size: 11px;
                border: 1px solid #444; border-radius: 4px;
                padding: 4px; margin-top: 2px;
            }
            QGroupBox::title { padding: 0 6px; }
            QGroupBox QLabel { font-size: 11px; }
            QGroupBox QCheckBox { font-size: 11px; }
        """)
        action_layout = QGridLayout(action_group)
        action_layout.setSpacing(3)
        action_layout.setContentsMargins(4, 4, 4, 4)

        # 表头
        action_layout.addWidget(QLabel("信号类型"), 0, 0)
        action_layout.addWidget(QLabel("启用"), 0, 1)
        action_layout.addWidget(QLabel("应对措施"), 0, 2)
        action_layout.addWidget(QLabel("Telegram"), 0, 3)
        for col in range(4):
            lbl = action_layout.itemAtPosition(0, col)
            if lbl and lbl.widget():
                lbl.widget().setStyleSheet("color: #888; font-weight: bold;")

        self.action_widgets: Dict[str, Dict[str, QWidget]] = {}
        signal_types = [
            ('trend_breakout', '🚀 趋势突破'),
            ('deep_pullback', '📉 大幅回调'),
            ('stabilization_breakout', '🎯 企稳突破'),
            ('volume_surge', '📊 放量异动'),
            ('momentum_divergence', '⚡ 动量背离'),
        ]

        for i, (key, label) in enumerate(signal_types):
            row = i + 1
            lbl = QLabel(label)
            lbl.setStyleSheet("color: #ffd166; font-weight: bold;")
            action_layout.addWidget(lbl, row, 0)

            enable_cb = QCheckBox()
            enable_cb.setChecked(True)
            enable_cb.setStyleSheet("QCheckBox::indicator { width: 16px; height: 16px; }")
            action_layout.addWidget(enable_cb, row, 1, alignment=Qt.AlignCenter)

            action_cb = QComboBox()
            action_cb.addItems(["仅提醒", "提醒+自动开仓", "提醒+自动平仓", "仅记录不提醒"])
            action_cb.setCurrentIndex(0)
            action_cb.setStyleSheet("""
                QComboBox {
                    background-color: #1e1e1e; color: #ffffff; font-size: 11px;
                    border: 1px solid #444; padding: 1px 4px;
                }
                QComboBox QAbstractItemView {
                    background-color: #1e1e1e; color: #dddddd; font-size: 11px;
                    selection-background-color: #2b5d87;
                }
            """)
            action_layout.addWidget(action_cb, row, 2)

            tg_cb = QCheckBox()
            tg_cb.setChecked(True)
            tg_cb.setStyleSheet("QCheckBox::indicator { width: 16px; height: 16px; }")
            action_layout.addWidget(tg_cb, row, 3, alignment=Qt.AlignCenter)

            self.action_widgets[key] = {
                'enable': enable_cb,
                'action': action_cb,
                'telegram': tg_cb,
            }

        # 全局 Telegram 总开关快捷行
        action_layout.addWidget(QLabel(""), len(signal_types) + 1, 0)
        global_tg_btn = QPushButton("全部Telegram开")
        global_tg_btn.setStyleSheet("""
            QPushButton { background-color: #3a3a3a; color: #ccc; padding: 3px 10px; border-radius: 3px; border: 1px solid #444; }
            QPushButton:hover { background-color: #4a4a4a; color: #fff; }
        """)
        global_tg_btn.clicked.connect(lambda: self._set_all_tg(True))
        action_layout.addWidget(global_tg_btn, len(signal_types) + 1, 1)

        global_tg_off_btn = QPushButton("全部Telegram关")
        global_tg_off_btn.setStyleSheet("""
            QPushButton { background-color: #3a3a3a; color: #ccc; padding: 3px 10px; border-radius: 3px; border: 1px solid #444; }
            QPushButton:hover { background-color: #4a4a4a; color: #fff; }
        """)
        global_tg_off_btn.clicked.connect(lambda: self._set_all_tg(False))
        action_layout.addWidget(global_tg_off_btn, len(signal_types) + 1, 2)

        bottom_layout.addWidget(action_group)

        self.bottom_scroll.setWidget(bottom_widget)

        # ── 外层分割器：可拖动控制 监控区/参数区 高度 ──
        outer_splitter = QSplitter(Qt.Vertical)
        outer_splitter.setHandleWidth(10)
        outer_splitter.setChildrenCollapsible(False)
        outer_splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #3a3a3a;
                border-top: 1px solid #555;
                border-bottom: 1px solid #555;
            }
            QSplitter::handle:hover {
                background-color: #ff6b35;
                border-top: 1px solid #ff6b35;
                border-bottom: 1px solid #ff6b35;
            }
            QSplitter::handle:pressed {
                background-color: #ff4500;
            }
        """)
        outer_splitter.addWidget(splitter)
        outer_splitter.addWidget(self.bottom_scroll)
        outer_splitter.setStretchFactor(0, 5)  # 监控区分更多空间
        outer_splitter.setStretchFactor(1, 2)  # 参数区
        outer_splitter.setSizes([600, 220])
        outer_layout.addWidget(outer_splitter, 10)

        # ═══ 定时推送到 Telegram ═══
        push_group = QGroupBox("📤 定时推送监控报表到 Telegram")
        push_group.setStyleSheet("""
            QGroupBox {
                color: #8fd3ff; font-weight: bold; font-size: 11px;
                border: 1px solid #3a5a7a; border-radius: 4px;
                padding: 4px; margin-top: 2px;
            }
            QGroupBox::title { padding: 0 6px; }
        """)
        push_layout = QHBoxLayout(push_group)
        push_layout.setSpacing(6)

        push_layout.addWidget(QLabel("推送间隔(分钟):"))
        self.push_interval_spin = QSpinBox()
        self.push_interval_spin.setRange(1, 1440)
        self.push_interval_spin.setValue(30)
        self.push_interval_spin.setSuffix(" 分钟")
        self.push_interval_spin.setStyleSheet("background-color: #1e1e1e; color: #fff; border: 1px solid #444; padding: 2px 4px; font-size: 11px;")
        push_layout.addWidget(self.push_interval_spin)

        push_layout.addWidget(QLabel("Bot Token:"))
        self.push_token_edit = QLineEdit()
        self.push_token_edit.setPlaceholderText("Bot Token")
        self.push_token_edit.setEchoMode(QLineEdit.Password)
        self.push_token_edit.setStyleSheet("background-color: #1e1e1e; color: #fff; border: 1px solid #444; padding: 3px 6px; font-size: 11px;")
        push_layout.addWidget(self.push_token_edit)

        push_layout.addWidget(QLabel("Chat ID:"))
        self.push_chat_id_edit = QLineEdit()
        self.push_chat_id_edit.setPlaceholderText("Chat ID")
        self.push_chat_id_edit.setStyleSheet("background-color: #1e1e1e; color: #fff; border: 1px solid #444; padding: 3px 6px; font-size: 11px;")
        push_layout.addWidget(self.push_chat_id_edit)

        self.push_start_btn = QPushButton("▶ 启动推送")
        self.push_start_btn.setStyleSheet("""
            QPushButton { background-color: #28a745; color: white; font-weight: bold; padding: 4px 12px; border-radius: 3px; font-size: 11px; border: 1px solid rgba(255,255,255,0.1); }
            QPushButton:hover { background-color: #34d058; }
        """)
        self.push_start_btn.clicked.connect(self.start_monitor_push)
        push_layout.addWidget(self.push_start_btn)

        self.push_stop_btn = QPushButton("⏹ 停止")
        self.push_stop_btn.setStyleSheet("""
            QPushButton { background-color: #6c4444; color: #ffaaaa; font-weight: bold; padding: 4px 12px; border-radius: 3px; font-size: 11px; border: 1px solid rgba(255,255,255,0.1); }
            QPushButton:hover { background-color: #885555; color: #ffcccc; }
        """)
        self.push_stop_btn.clicked.connect(self.stop_monitor_push)
        push_layout.addWidget(self.push_stop_btn)

        self.push_status_label = QLabel("⏸ 未启动")
        self.push_status_label.setStyleSheet("color: #aaaaaa; font-size: 11px;")
        push_layout.addWidget(self.push_status_label)

        outer_layout.addWidget(push_group)

        # ── 初始化推送定时器 ──
        self._push_timer = QTimer(self)
        self._push_timer.timeout.connect(self._do_monitor_push)
        self._push_config_path = self._config_dir / "monitor_push_config.json"
        self._load_push_config()

    def _get_slider_value(self, key: str) -> float:
        """返回滑轨的真实值（整数值 / scale）"""
        s = self._sliders.get(key)
        return s['slider'].value() / s['scale'] if s else 0.0

    def _set_slider_value(self, key: str, value) -> None:
        """设置滑轨值并刷新显示标签"""
        s = self._sliders.get(key)
        if not s:
            return
        s['slider'].blockSignals(True)
        s['slider'].setValue(int(round(float(value) * s['scale'])))
        s['slider'].blockSignals(False)
        v = s['slider'].value()
        rv = v / s['scale']
        fmt = s['fmt']
        disp = f"{int(rv):{fmt}}" if fmt == 'd' else f"{rv:{fmt}}"
        s['label'].setText(f"{disp} {s['unit']}")

    def _make_spin(self, lo: int, hi: int, val: int, suffix: str) -> QSpinBox:
        spin = QSpinBox()
        spin.setRange(lo, hi)
        spin.setValue(val)
        spin.setSuffix(f" {suffix}")
        spin.setStyleSheet("background-color: #1e1e1e; color: #ffffff; border: 1px solid #444; padding: 1px 2px; font-size: 11px;")
        return spin

    def _make_double_spin(self, lo: float, hi: float, val: float, step: float, suffix: str) -> QDoubleSpinBox:
        spin = QDoubleSpinBox()
        spin.setRange(lo, hi)
        spin.setValue(val)
        spin.setSingleStep(step)
        spin.setDecimals(1)
        spin.setSuffix(f" {suffix}")
        spin.setStyleSheet("background-color: #1e1e1e; color: #ffffff; border: 1px solid #444; padding: 1px 2px; font-size: 11px;")
        return spin

    # ── 风格预设 ──
    PRESETS = {
        "conservative": {
            "check_interval_sec": 120,
            "trend_breakout_pct": 3.5,
            "trend_breakout_volume_ratio": 2.0,
            "deep_pullback_pct": 7.0,
            "stabilization_max_range_pct": 0.6,
            "stabilization_breakout_volume_ratio": 2.2,
            "volume_surge_ratio": 4.0,
            "min_signal_score": 70,
            "signal_cooldown_min": 30,
            "description": "高阈值 · 低频率 · 适合 BTC/ETH 等主流币",
        },
        "neutral": {
            "check_interval_sec": 60,
            "trend_breakout_pct": 2.5,
            "trend_breakout_volume_ratio": 1.5,
            "deep_pullback_pct": 5.0,
            "stabilization_max_range_pct": 1.2,
            "stabilization_breakout_volume_ratio": 1.8,
            "volume_surge_ratio": 3.0,
            "min_signal_score": 60,
            "signal_cooldown_min": 15,
            "description": "均衡配置 · 适中频率 · 通用推荐",
        },
        "aggressive": {
            "check_interval_sec": 30,
            "trend_breakout_pct": 1.5,
            "trend_breakout_volume_ratio": 1.2,
            "deep_pullback_pct": 3.0,
            "stabilization_max_range_pct": 2.0,
            "stabilization_breakout_volume_ratio": 1.3,
            "volume_surge_ratio": 2.0,
            "min_signal_score": 45,
            "signal_cooldown_min": 8,
            "description": "低阈值 · 高频率 · 适合波动大的山寨币",
        },
    }

    PRESET_COLORS = {
        "conservative": "#2d6a4f",
        "neutral": "#3a6b8f",
        "aggressive": "#b8860b",
    }

    def _apply_preset(self, preset_name: str):
        """一键应用风格预设"""
        cfg = self.PRESETS.get(preset_name)
        if not cfg:
            return

        for key in [
            'check_interval_sec', 'trend_breakout_pct', 'trend_breakout_volume_ratio',
            'deep_pullback_pct', 'stabilization_max_range_pct',
            'stabilization_breakout_volume_ratio', 'volume_surge_ratio',
            'min_signal_score', 'signal_cooldown_min',
        ]:
            if key in cfg:
                self._set_slider_value(key, cfg[key])

        # 高亮当前选中的按钮
        self._active_preset = preset_name
        for key, btn in self._preset_buttons.items():
            color = self.PRESET_COLORS.get(key, "#444")
            if key == preset_name:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color}; color: white; font-weight: bold;
                        padding: 2px 8px; border-radius: 3px; font-size: 10px;
                        border: 2px solid white;
                    }}
                    QPushButton:hover {{ border: 2px solid white; }}
                """)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color}; color: white; font-weight: bold;
                        padding: 2px 8px; border-radius: 3px; font-size: 10px;
                        border: 1px solid rgba(255,255,255,0.15);
                    }}
                    QPushButton:hover {{ border: 1px solid rgba(255,255,255,0.4); }}
                """)

        self.on_log_message(f"风格预设已切换为: {preset_name} — {cfg['description']}", "INFO")

    def _quick_add(self, inst_id: str):
        """快捷添加交易对"""
        self.add_input.setCurrentText(inst_id)
        self.add_pair_from_input()

    def _set_trend_cell(self, row: int, inst_id: str):
        """从缓存设置走势预判列（显示日/时/分三周期）"""
        trend = self._pair_trend_cache.get(inst_id)
        if trend:
            d = trend.get("1D", ("--", "#aaaaaa"))
            h = trend.get("1H", ("--", "#aaaaaa"))
            m = trend.get("3m", ("--", "#aaaaaa"))
            text = f"D:{d[0]} | H:{h[0]} | M:{m[0]}"
            item = QTableWidgetItem(text)
            # 取日线颜色作为主色
            item.setForeground(QColor(d[1]))
            self.pair_table.setItem(row, 7, item)
        else:
            self.pair_table.setItem(row, 7, QTableWidgetItem("--"))

    def _set_all_tg(self, checked: bool):
        """全部 Telegram 开/关"""
        for w in self.action_widgets.values():
            w['telegram'].setChecked(checked)

    def _get_signal_action_key(self, signal_type: str) -> Optional[str]:
        """将信号类型中文名映射到 action_widgets 的 key"""
        _map = {
            '趋势突破': 'trend_breakout',
            '大幅回调': 'deep_pullback',
            '企稳突破': 'stabilization_breakout',
            '放量异动': 'volume_surge',
            '动量背离': 'momentum_divergence',
        }
        return _map.get(signal_type)

    def _get_signal_action(self, signal_type: str) -> Dict:
        """获取某个信号类型的应对措施配置"""
        key = self._get_signal_action_key(signal_type)
        if key and key in self.action_widgets:
            w = self.action_widgets[key]
            return {
                'enable': w['enable'].isChecked(),
                'action': w['action'].currentText(),
                'telegram': w['telegram'].isChecked(),
            }
        return {'enable': True, 'action': '仅提醒', 'telegram': True}

    def _open_signal_log_window(self):
        """打开独立的信号日志弹窗，并回放已有日志"""
        if self._signal_log_window is None:
            self._signal_log_window = SignalLogWindow(self)
            self._signal_log_window.finished.connect(self._on_signal_log_window_closed)
            # 回放已有日志内容
            existing_html = self.signal_log.toHtml()
            if existing_html:
                self._signal_log_window.log_text.setHtml(existing_html)
                scrollbar = self._signal_log_window.log_text.verticalScrollBar()
                scrollbar.setValue(scrollbar.maximum())
            self._signal_log_window._signal_count = len(self._signals_log)
            self._signal_log_window.win_count_label.setText(f"{len(self._signals_log)} 条信号")
            if self.monitor_worker and self.monitor_worker.isRunning():
                self._signal_log_window.set_status("🟢 监控中", "#00ff88")
            else:
                self._signal_log_window.set_status("⏸ 已停止", "#aaaaaa")
        self._signal_log_window.show()
        self._signal_log_window.raise_()
        self._signal_log_window.activateWindow()

    def _on_signal_log_window_closed(self):
        """弹窗关闭回调"""
        self._signal_log_window = None

    def _write_to_log_window(self, html: str, is_signal: bool = False):
        """同时写入内嵌日志和独立窗口"""
        self.signal_log.append(html)
        if self._signal_log_window and self._signal_log_window.isVisible():
            self._signal_log_window.append_log(html)
            if is_signal:
                self._signal_log_window.increment_count()

    def _load_swap_pairs(self):
        """从 OKX 异步加载所有 USDT 永续合约交易对 (后台线程)"""
        if not self.okx_client:
            return

        # 使用 Python 线程避免阻塞 UI
        import threading

        def _fetch():
            try:
                result = self.okx_client.get_tickers("SWAP")
                if result.get('code') != '0':
                    # 通过信号报告错误
                    QTimer.singleShot(0, lambda: self.on_log_message(
                        f"加载交易对失败: {result.get('msg', '未知错误')}", "WARNING"))
                    return
                all_pairs = []
                for item in result.get('data', []):
                    inst_id = item.get('instId', '')
                    if inst_id.endswith('-USDT-SWAP'):
                        all_pairs.append(inst_id)
                all_pairs.sort()
                # 回到主线程更新 UI
                QTimer.singleShot(0, lambda: self._on_swap_pairs_loaded(all_pairs))
            except Exception as e:
                QTimer.singleShot(0, lambda: self.on_log_message(
                    f"加载交易对列表异常: {e}", "ERROR"))

        threading.Thread(target=_fetch, daemon=True).start()

    def _on_swap_pairs_loaded(self, all_pairs: List[str]):
        """主线程回调：更新下拉框"""
        self._swap_pairs = all_pairs
        self.add_input.clear()
        self.add_input.addItems(all_pairs)
        self._pair_completer.model().setStringList(all_pairs)
        self.on_log_message(f"已加载 {len(all_pairs)} 个 USDT 永续合约交易对", "INFO")

    def add_pair_from_input(self):
        """从下拉框/输入框添加交易对"""
        text = self.add_input.currentText().strip()
        if not text:
            return

        pairs = [text.upper()]  # 单次添加一个
        # 确保格式
        formatted = []
        for p in pairs:
            p = p.strip().upper()
            if not p.endswith('-USDT-SWAP'):
                if p.endswith('-USDT'):
                    p = p.replace('-USDT', '-USDT-SWAP')
                elif '/USDT' in p:
                    p = p.replace('/USDT', '-USDT-SWAP')
                elif '-SWAP' not in p:
                    p = f"{p}-USDT-SWAP"
            formatted.append(p)

        if self.monitor_worker and self.monitor_worker.isRunning():
            for p in formatted:
                self.monitor_worker.add_pair(p)
            self.refresh_pair_table()
            self.save_config()
        else:
            # 未启动监控时，直接添加到表格
            for p in formatted:
                # 检查是否重复
                existing = self._get_current_pairs()
                if p not in existing:
                    row = self.pair_table.rowCount()
                    self.pair_table.insertRow(row)
                    self.pair_table.setItem(row, 0, QTableWidgetItem(p))
                    self.pair_table.setItem(row, 1, QTableWidgetItem("--"))
                    self.pair_table.setItem(row, 2, QTableWidgetItem("--"))
                    self.pair_table.setItem(row, 3, QTableWidgetItem("--"))
                    self.pair_table.setItem(row, 4, QTableWidgetItem("--"))
                    self.pair_table.setItem(row, 5, QTableWidgetItem("待启动"))
                    self.pair_table.setItem(row, 6, QTableWidgetItem("--"))
                    self._set_trend_cell(row, p)
            self.pair_count_label.setText(f"{self.pair_table.rowCount()} 个交易对")
            self.save_config()

    def clear_signal_log(self):
        """清空信号日志"""
        self._signals_log.clear()
        self.signal_log.clear()
        self.signal_count_label.setText("0 条信号")
        if self._signal_log_window and self._signal_log_window.isVisible():
            self._signal_log_window.clear_log()

    def _forward_to_telegram(self, sig_dict: Dict, action_cfg: Dict = None):
        """转发信号到 Telegram，包含应对措施"""
        emoji_map = {
            '趋势突破': '🚀',
            '大幅回调': '📉',
            '企稳突破': '🎯',
            '放量异动': '📊',
            '动量背离': '⚡',
        }
        emoji = emoji_map.get(sig_dict.get('signal_type', ''), '📡')
        dire_emoji = '🟢' if sig_dict.get('direction') == 'BUY' else ('🔴' if sig_dict.get('direction') == 'SHORT' else '⚪')

        action_text = ''
        if action_cfg:
            action_text = f"\n⚡ 应对措施: **{action_cfg.get('action', '仅提醒')}**"

        msg = (
            f"{emoji} **{sig_dict.get('signal_type', '')}** {dire_emoji}\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📊 交易对: `{sig_dict.get('inst_id', '')}`\n"
            f"🎯 方向: **{sig_dict.get('direction', '')}**\n"
            f"⭐ 评分: **{sig_dict.get('score', 0):.0f}** 分\n"
            f"💰 价格: ${sig_dict.get('price', 0):.4f}\n"
            f"📝 {sig_dict.get('message', '')}\n"
            f"🕐 {sig_dict.get('timestamp', '')}"
            f"{action_text}"
        )
        self.telegram_alert_requested.emit("monitor_pool", msg)

    def apply_settings(self):
        """应用监测参数"""
        if self.monitor_worker:
            config = self._collect_config()
            self.monitor_worker.update_config(config)
            self.on_log_message("监测参数已更新", "SUCCESS")

    def _collect_config(self) -> Dict:
        config = {
            'check_interval_sec': int(self._get_slider_value('check_interval_sec')),
            'trend_breakout_pct': self._get_slider_value('trend_breakout_pct'),
            'trend_breakout_volume_ratio': self._get_slider_value('trend_breakout_volume_ratio'),
            'deep_pullback_pct': self._get_slider_value('deep_pullback_pct'),
            'stabilization_max_range_pct': self._get_slider_value('stabilization_max_range_pct'),
            'stabilization_breakout_volume_ratio': self._get_slider_value('stabilization_breakout_volume_ratio'),
            'volume_surge_ratio': self._get_slider_value('volume_surge_ratio'),
            'min_signal_score': int(self._get_slider_value('min_signal_score')),
            'signal_cooldown_min': int(self._get_slider_value('signal_cooldown_min')),
        }
        # 添加应对措施配置
        config['signal_actions'] = {}
        for key, widgets in self.action_widgets.items():
            config['signal_actions'][key] = {
                'enable': widgets['enable'].isChecked(),
                'action': widgets['action'].currentText(),
                'telegram': widgets['telegram'].isChecked(),
            }
        return config

    def _get_current_pairs(self) -> List[str]:
        """获取当前表格中的交易对列表"""
        pairs = []
        for row in range(self.pair_table.rowCount()):
            item = self.pair_table.item(row, 0)
            if item:
                pairs.append(item.text())
        return pairs

    def load_pairs_from_config(self) -> List[str]:
        """从配置文件加载交易对列表"""
        if self._pool_config_path.exists():
            try:
                with open(self._pool_config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                return cfg.get('pairs', [])
            except Exception:
                pass
        return []

    def _on_column_resized(self, logical_index: int, old_size: int, new_size: int):
        """列宽拖拽变化时自动保存（200ms 防抖）"""
        if self._col_resize_timer is None:
            self._col_resize_timer = QTimer(self)
            self._col_resize_timer.setSingleShot(True)
            self._col_resize_timer.timeout.connect(self._save_column_widths)
        else:
            self._col_resize_timer.stop()
        self._col_resize_timer.start(200)

    def _save_column_widths(self):
        """仅保存列宽到配置文件"""
        try:
            if self._pool_config_path.exists():
                with open(self._pool_config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
            else:
                cfg = {'pairs': self._get_current_pairs(), 'pinned': list(self._pinned_pairs), 'settings': self._collect_config()}
            cfg['col_widths'] = [self.pair_table.columnWidth(i) for i in range(self.pair_table.columnCount())]
            cfg['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(self._pool_config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def save_config(self, pairs: List[str] = None):
        """保存监控池配置"""
        if pairs is None:
            pairs = self._get_current_pairs()
        try:
            col_widths = [self.pair_table.columnWidth(i) for i in range(self.pair_table.columnCount())]
            cfg = {
                'pairs': pairs,
                'pinned': list(self._pinned_pairs),
                'col_widths': col_widths,
                'settings': self._collect_config(),
                'updated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            with open(self._pool_config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"保存监控池配置失败: {e}")

    def load_config(self):
        """加载配置并恢复"""
        pairs = []
        if self._pool_config_path.exists():
            try:
                with open(self._pool_config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                pairs = cfg.get('pairs', [])
                settings = cfg.get('settings', {})
                self._pinned_pairs = set(cfg.get('pinned', []))
                self._saved_col_widths = cfg.get('col_widths', [])
                self._apply_settings_to_ui(settings)
            except Exception:
                pass

        # 填充表格（置顶优先）
        pairs = sorted(pairs, key=lambda p: (0 if p in self._pinned_pairs else 1, p))
        for inst_id in pairs:
            row = self.pair_table.rowCount()
            self.pair_table.insertRow(row)
            self.pair_table.setItem(row, 0, QTableWidgetItem(inst_id))
            if inst_id in self._pinned_pairs:
                pin_item = self.pair_table.item(row, 0)
                if pin_item:
                    pin_item.setForeground(QColor("#ffd700"))
            self.pair_table.setItem(row, 1, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 2, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 3, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 4, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 5, QTableWidgetItem("待启动"))
            self.pair_table.setItem(row, 6, QTableWidgetItem("--"))
            self._set_trend_cell(row, inst_id)

        self.pair_count_label.setText(f"{len(pairs)} 个交易对")
        # 恢复保存的列宽
        saved_w = getattr(self, '_saved_col_widths', [])
        if saved_w and len(saved_w) == self.pair_table.columnCount():
            for i, w in enumerate(saved_w):
                self.pair_table.setColumnWidth(i, max(30, int(w)))

    def _apply_settings_to_ui(self, settings: Dict):
        if not settings:
            return
        for key in [
            'check_interval_sec', 'trend_breakout_pct', 'trend_breakout_volume_ratio',
            'deep_pullback_pct', 'stabilization_max_range_pct',
            'stabilization_breakout_volume_ratio', 'volume_surge_ratio',
            'min_signal_score', 'signal_cooldown_min',
        ]:
            if key in settings:
                try:
                    self._set_slider_value(key, settings[key])
                except Exception:
                    pass

        # 恢复应对措施配置
        signal_actions = settings.get('signal_actions', {})
        for key, action_cfg in signal_actions.items():
            if key in self.action_widgets:
                w = self.action_widgets[key]
                try:
                    w['enable'].setChecked(action_cfg.get('enable', True))
                except Exception:
                    pass
                try:
                    idx = w['action'].findText(action_cfg.get('action', '仅提醒'))
                    if idx >= 0:
                        w['action'].setCurrentIndex(idx)
                except Exception:
                    pass
                try:
                    w['telegram'].setChecked(action_cfg.get('telegram', True))
                except Exception:
                    pass

    def remove_selected_pair(self):
        """移除选中的交易对"""
        rows = set()
        for idx in self.pair_table.selectedIndexes():
            rows.add(idx.row())
        if not rows:
            return
        if self.monitor_worker and self.monitor_worker.isRunning():
            for row in sorted(rows, reverse=True):
                item = self.pair_table.item(row, 0)
                if item:
                    self.monitor_worker.remove_pair(item.text())
        elif not self.monitor_worker or not self.monitor_worker.isRunning():
            for row in sorted(rows, reverse=True):
                self.pair_table.removeRow(row)
            self.pair_count_label.setText(f"{self.pair_table.rowCount()} 个交易对")
            self.save_config()
            return
        self.refresh_pair_table()

    def set_trade_runtime_dependencies(self, trade_executor=None, trade_settings_manager=None):
        """注入自动交易依赖。"""
        if trade_executor is not None:
            self.trade_executor = trade_executor
        if trade_settings_manager is not None:
            self.trade_settings_manager = trade_settings_manager
        self.save_config()

    def set_scan_auto_trader(self, trader):
        """
        注入扫描驱动自动交易编排器引用，用于桥接监控池高质量信号。

        当监控池发现评分 ≥ high_signal_threshold 的信号时，通过
        MonitorWorker.high_quality_signal → ScanDrivenAutoTrader.on_monitor_signal
        路径自动触发入场评估。

        trader=None 时断开连接（编排器停止时调用）。
        """
        self._scan_auto_trader_ref = trader
        # 若监控工作线程已在运行，立即连接/断开
        if self.monitor_worker and self.monitor_worker.isRunning():
            try:
                if trader is not None:
                    self.monitor_worker.high_quality_signal.connect(
                        trader.on_monitor_signal, Qt.QueuedConnection
                    )
                else:
                    # 无法直接断开（不知道之前连接的是哪个 trader），
                    # 线程重启时会重新建立连接，此处仅清除引用
                    pass
            except Exception:
                pass

    def show_pair_menu(self, pos):
        """右键菜单"""
        # 右键点击的行加入选中（不覆盖已有选中）
        clicked_row = self.pair_table.rowAt(pos.y())
        if clicked_row >= 0:
            if not self.pair_table.selectionModel().isRowSelected(clicked_row):
                # Ctrl+click behavior: 追加选中而不清除已有
                self.pair_table.selectionModel().select(
                    self.pair_table.model().index(clicked_row, 0),
                    self.pair_table.selectionModel().Select | self.pair_table.selectionModel().Rows
                )

        selected_rows = set(idx.row() for idx in self.pair_table.selectedIndexes())
        has_pinned = any(
            self.pair_table.item(r, 0) and self.pair_table.item(r, 0).text() in self._pinned_pairs
            for r in selected_rows
        )
        all_pinned = all(
            self.pair_table.item(r, 0) and self.pair_table.item(r, 0).text() in self._pinned_pairs
            for r in selected_rows
        ) if selected_rows else False

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: #2a2a2a; color: #dddddd; border: 1px solid #444; }
            QMenu::item:selected { background-color: #3a3a3a; }
        """)
        if all_pinned:
            pin_label = f"📌 取消置顶 ({len(selected_rows)}个)"
        elif has_pinned:
            pin_label = f"📌 全部置顶 ({len(selected_rows)}个)"
        else:
            pin_label = f"📌 置顶 ({len(selected_rows)}个)" if len(selected_rows) > 1 else "📌 置顶 / 取消置顶"
        pin_action = menu.addAction(pin_label)
        pin_action.triggered.connect(self._toggle_pin_pair)
        menu.addSeparator()
        remove_action = menu.addAction("🗑 从监控池移除")
        remove_action.triggered.connect(self.remove_selected_pair)
        menu.exec(self.pair_table.viewport().mapToGlobal(pos))

    def _toggle_pin_pair(self):
        """切换选中交易对的置顶状态（支持多选）"""
        rows = set(idx.row() for idx in self.pair_table.selectedIndexes())
        if not rows:
            return
        # 检查当前状态：全置顶则取消，否则全部置顶
        all_pinned = all(
            self.pair_table.item(r, 0) and self.pair_table.item(r, 0).text() in self._pinned_pairs
            for r in rows
        )
        count = 0
        for row in rows:
            item = self.pair_table.item(row, 0)
            if not item:
                continue
            inst_id = item.text()
            if all_pinned:
                if inst_id in self._pinned_pairs:
                    self._pinned_pairs.discard(inst_id)
                    count += 1
            else:
                if inst_id not in self._pinned_pairs:
                    self._pinned_pairs.add(inst_id)
                    count += 1
        if count > 0:
            self.on_log_message(f"📌 {count}个交易对{'取消置顶' if all_pinned else '已置顶'}", "SUCCESS")
        self.refresh_pair_table()
        self.save_config()

    def toggle_monitor(self):
        """启动/停止监控"""
        if self.monitor_worker and self.monitor_worker.isRunning():
            self.monitor_worker.stop()
            self.monitor_worker.wait(200)   # 非阻塞：最多等 200ms，不卡 UI
            self.start_btn.setText("▶ 启动监控")
            self.start_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745; color: white; font-weight: bold;
                    padding: 6px 16px; border-radius: 4px;
                }
                QPushButton:hover { background-color: #34d058; }
            """)
            self.status_label.setText("⏸ 已停止")
            self.status_label.setStyleSheet("color: #aaaaaa;")
            if self._signal_log_window and self._signal_log_window.isVisible():
                self._signal_log_window.set_status("⏸ 已停止", "#aaaaaa")
        else:
            if not self.okx_client:
                QMessageBox.warning(self, "错误", "未连接OKX客户端，无法启动监控")
                return
            self._start_monitor()

    def _start_monitor(self):
        """启动监控工作线程"""
        config = self._collect_config()
        pairs = self._get_current_pairs()
        self.monitor_worker = MonitorWorker(self.okx_client, config, pairs)
        self.monitor_worker.signal_detected.connect(self.on_signal_detected, Qt.QueuedConnection)
        self.monitor_worker.status_update.connect(self.on_status_update, Qt.QueuedConnection)
        self.monitor_worker.log_message.connect(self.on_log_message, Qt.QueuedConnection)
        self.monitor_worker.pair_data_updated.connect(self.on_pair_data_updated, Qt.QueuedConnection)
        # ── 桥接高质量信号到扫描驱动自动交易编排器 ─────────────────────────────
        _trader = getattr(self, '_scan_auto_trader_ref', None)
        if _trader is not None:
            self.monitor_worker.high_quality_signal.connect(
                _trader.on_monitor_signal, Qt.QueuedConnection
            )
        self.monitor_worker.start()
        self.start_btn.setText("⏹ 停止监控")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545; color: white; font-weight: bold;
                padding: 6px 16px; border-radius: 4px;
            }
            QPushButton:hover { background-color: #e8445a; }
        """)
        self.status_label.setText("🟢 监控中")
        self.status_label.setStyleSheet("color: #00ff88; font-size: 12px; font-weight: bold;")
        if self._signal_log_window and self._signal_log_window.isVisible():
            self._signal_log_window.set_status("🟢 监控中", "#00ff88")

    def on_signal_detected(self, signal):
        """处理检测到的信号"""
        sig_dict = signal.to_dict()
        self._signals_log.insert(0, sig_dict)
        if len(self._signals_log) > self._max_log_entries:
            self._signals_log = self._signals_log[:self._max_log_entries]
        color_map = {'BUY': '#00ff88', 'SHORT': '#ff6666', 'NEUTRAL': '#ffaa00'}
        dire_color = color_map.get(signal.direction, '#ffffff')
        log_entry = (
            f"<span style='color:#888'>[{signal.timestamp}]</span> "
            f"<span style='color:#ff6b35'>{signal.signal_type}</span> "
            f"<b style='color:{dire_color}'>{signal.direction}</b> "
            f"<span style='color:#ffd166'>{signal.inst_id}</span> "
            f"({signal.score:.0f}分) "
            f"<span style='color:#aaa'>{signal.message}</span>"
        )
        self._write_to_log_window(log_entry, is_signal=True)
        suggestion = TradingSuggester.generate(signal)
        advice = (
            f"<span style='color:#666'> ⤷ 建议: </span>"
            f"<b style='color:{suggestion['action_color']}'>{suggestion['action']}</b>"
            f"<span style='color:#666'> | 走势: </span>"
            f"<span style='color:{suggestion['bias_color']}'>{suggestion['trend_phase']}·{suggestion['bias']}</span>"
            f"<br><span style='color:#777;font-size:10px'>     {suggestion['summary']}</span>"
        )
        self._write_to_log_window(advice)
        scrollbar = self.signal_log.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
        self.signal_count_label.setText(f"{len(self._signals_log)} 条信号")
        # 融合实时指标推断走势（live_data 提供 MACD/BB/D1/RSI 真实数据）
        live = self._pair_live_data.get(signal.inst_id, {})
        self._pair_trend_cache[signal.inst_id] = MultiTimeframeAnalyzer.analyze(signal, live)
        action_cfg = self._get_signal_action(signal.signal_type)
        self._execute_signal_action(signal, action_cfg)

        # ── 弹窗提醒（非模态，自动30秒消失）──
        if action_cfg['enable'] and action_cfg['action'] != '仅记录不提醒':
            alert = SignalAlertDialog(signal, suggestion, self)
            alert.show()

        if self.telegram_checkbox.isChecked() and action_cfg['enable'] and action_cfg['telegram']:
            self._forward_to_telegram(sig_dict, action_cfg)
        self.refresh_pair_table()

    def add_pair(self, inst_id: str):
        """供外部页面直接加入监控池。"""
        if not inst_id:
            return
        self.add_input.setCurrentText(inst_id)
        self.add_pair_from_input()

    def add_pair_with_action_preset(self, inst_id: str, signal_type: str, action_text: str = "提醒+自动开仓"):
        """加入监控池并预设该信号类型的应对措施。"""
        self.add_pair(inst_id)
        key = self._get_signal_action_key(signal_type)
        if not key or key not in self.action_widgets:
            self.on_log_message(f"{inst_id} 已加入监控池，但未识别到可预设的信号类型：{signal_type or '-'}", "WARNING")
            return
        widgets = self.action_widgets[key]
        widgets['enable'].setChecked(True)
        idx = widgets['action'].findText(action_text)
        if idx >= 0:
            widgets['action'].setCurrentIndex(idx)
        self.save_config()
        self.on_log_message(f"{inst_id} 已加入监控池，并将“{signal_type}”预设为“{action_text}”", "SUCCESS")

    def _shared_auto_settings(self) -> Dict:
        if not self.trade_settings_manager:
            return {
                'position_size': 0.10,
                'leverage': 3,
                'take_profit_pct': 5.0,
                'stop_loss_pct': 3.0,
                'allow_short': True,
                'prefer_market_order': True,
                'signal_cooldown_minutes': 15,
            }
        settings = self.trade_settings_manager.get_all()
        merged = {}
        merged.update(settings.get('common', {}))
        merged.update(settings.get('auto_trading', {}))
        return merged

    def _execute_signal_action(self, signal, action_cfg: Dict):
        """按监控池动作配置执行共享风控下单/平仓。"""
        if not action_cfg.get('enable'):
            return
        action_name = str(action_cfg.get('action') or '')
        if action_name not in {"提醒+自动开仓", "提醒+自动平仓"}:
            return
        if not self.trade_executor:
            self.on_log_message("监控池已配置自动执行，但当前未注入交易执行器，已跳过。", "WARNING")
            return

        settings = self._shared_auto_settings()
        inst_id = str(getattr(signal, 'inst_id', '') or '').strip().upper()
        direction = str(getattr(signal, 'direction', '') or '').upper()
        if not inst_id:
            return

        # 防止与扫描器 auto-trader 重复开仓（同 symbol 冷却锁）
        cooldown_key = f"monitor_auto:{inst_id}"
        now_ts = time.time()
        last_exec = float(self._auto_trade_cooldowns.get(cooldown_key, 0.0) or 0.0)
        if now_ts - last_exec < 15:  # 15秒内不重复触发
            return

        if action_name == "提醒+自动开仓":
            self._auto_trade_cooldowns[cooldown_key] = now_ts
            self._auto_open_from_signal(inst_id, direction, signal, settings)
        elif action_name == "提醒+自动平仓":
            self._auto_trade_cooldowns[cooldown_key] = now_ts
            self._auto_close_from_signal(inst_id, signal)

    def _auto_open_from_signal(self, inst_id: str, direction: str, signal, settings: Dict):
        if direction not in {"BUY", "LONG", "SHORT", "SELL"}:
            self.on_log_message(f"{inst_id} 自动开仓跳过：信号方向 {direction or '-'} 不支持自动执行", "WARNING")
            return
        if direction in {"SHORT", "SELL"} and not bool(settings.get('allow_short', True)):
            self.on_log_message(f"{inst_id} 自动开仓跳过：共享风控已关闭做空执行", "WARNING")
            return

        cooldown_minutes = int(settings.get('signal_cooldown_minutes', 15) or 0)
        cooldown_key = f"{inst_id}:{'SHORT' if direction in {'SHORT', 'SELL'} else 'LONG'}"
        last_ts = float(self._auto_trade_cooldowns.get(cooldown_key, 0.0) or 0.0)
        now_ts = time.time()
        if cooldown_minutes > 0 and (now_ts - last_ts) < cooldown_minutes * 60:
            remain = int(cooldown_minutes * 60 - (now_ts - last_ts))
            self.on_log_message(f"{inst_id} 自动开仓冷却中，剩余约 {remain}s", "INFO")
            return

        positions = self.trade_executor.get_positions(inst_id)
        current_pos = positions.get(inst_id)
        if current_pos:
            side_name = getattr(current_pos.side, 'name', str(current_pos.side)).upper()
            target_side = "SHORT" if direction in {"SHORT", "SELL"} else "LONG"
            if side_name == target_side:
                self.on_log_message(f"{inst_id} 已有同向持仓，跳过重复自动开仓", "INFO")
                return
            self.on_log_message(f"{inst_id} 当前存在反向/其他持仓，监控池不自动反手，已跳过", "WARNING")
            return

        try:
            balance = float(self.trade_executor.get_usdt_balance() or 0.0)
        except Exception as exc:
            self.on_log_message(f"{inst_id} 自动开仓前获取余额失败：{exc}", "ERROR")
            return
        position_size = float(settings.get('position_size', 0.10) or 0.10)
        usdt_amount = balance * position_size if position_size <= 1 else position_size
        if usdt_amount <= 0:
            self.on_log_message(f"{inst_id} 自动开仓跳过：可用资金不足", "WARNING")
            return

        order_type = "market" if bool(settings.get('prefer_market_order', True)) else "limit"
        limit_price = None if order_type == "market" else float(getattr(signal, 'price', 0.0) or 0.0)
        target_direction = "SHORT" if direction in {"SHORT", "SELL"} else "LONG"
        result = self.trade_executor.execute_entry(
            inst_id,
            target_direction,
            usdt_amount=usdt_amount,
            leverage=int(settings.get('leverage', 3) or 3),
            tp_pct=float(settings.get('take_profit_pct', 5.0) or 5.0) / 100.0,
            sl_pct=float(settings.get('stop_loss_pct', 3.0) or 3.0) / 100.0,
            order_type=order_type,
            price=limit_price,
        )
        if result.success:
            self._auto_trade_cooldowns[cooldown_key] = now_ts
            self.on_log_message(
                f"{inst_id} 监控池自动开仓成功：{target_direction}，投入约 {usdt_amount:.2f} USDT，"
                f"止盈 {float(settings.get('take_profit_pct', 5.0)):.2f}% / 止损 {float(settings.get('stop_loss_pct', 3.0)):.2f}%",
                "SUCCESS"
            )
        else:
            self.on_log_message(f"{inst_id} 监控池自动开仓失败：{result.message}", "ERROR")

    def _auto_close_from_signal(self, inst_id: str, signal):
        positions = self.trade_executor.get_positions(inst_id)
        if inst_id not in positions:
            self.on_log_message(f"{inst_id} 自动平仓跳过：当前无持仓", "INFO")
            return
        result = self.trade_executor.close_position(inst_id)
        if result.success:
            self.on_log_message(
                f"{inst_id} 监控池自动平仓成功：触发信号 {getattr(signal, 'signal_type', '-')}",
                "SUCCESS"
            )
        else:
            self.on_log_message(f"{inst_id} 监控池自动平仓失败：{result.message}", "ERROR")

    def on_status_update(self, inst_id: str, status: str):
        """更新交易对状态"""
        for row in range(self.pair_table.rowCount()):
            item = self.pair_table.item(row, 0)
            if item and item.text() == inst_id:
                self.pair_table.setItem(row, 5, QTableWidgetItem(status))
                break

    def on_log_message(self, message: str, level: str):
        """记录日志"""
        color = {'SUCCESS': '#00ff88', 'INFO': '#8fd3ff', 'WARNING': '#ffaa00', 'ERROR': '#ff6666'}.get(level, '#aaa')
        ts = datetime.now().strftime("%H:%M:%S")
        html = f"<span style='color:{color}'>[{ts}] [{level}] {message}</span>"
        self._write_to_log_window(html)
        scrollbar = self.signal_log.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def on_pair_data_updated(self, inst_id: str, data: dict):
        """更新交易对数据（价格/RSI/量比 + 走势预判刷新）"""
        # ── 缓存最新实时指标，供走势预判使用 ──
        self._pair_live_data[inst_id] = data

        # ── 无信号时也能从纯指标推断走势预判 ──
        if inst_id not in self._pair_trend_cache:
            self._pair_trend_cache[inst_id] = MultiTimeframeAnalyzer.from_live_data(data)

        for row in range(self.pair_table.rowCount()):
            item = self.pair_table.item(row, 0)
            if item and item.text() == inst_id:
                if 'price' in data:
                    self.pair_table.setItem(row, 1, QTableWidgetItem(f"${data['price']:.4f}"))
                if 'rsi' in data:
                    rsi = data['rsi']
                    rsi_item = QTableWidgetItem(f"{rsi:.1f}")
                    if rsi >= 70:
                        rsi_item.setForeground(QColor("#ff4444"))
                    elif rsi <= 30:
                        rsi_item.setForeground(QColor("#00ff88"))
                    else:
                        rsi_item.setForeground(QColor("#dddddd"))
                    self.pair_table.setItem(row, 3, rsi_item)
                if 'volume_ratio' in data:
                    vr = data['volume_ratio']
                    vr_item = QTableWidgetItem(f"{vr:.1f}x")
                    if vr >= 3.0:
                        vr_item.setForeground(QColor("#ffaa00"))
                    elif vr >= 2.0:
                        vr_item.setForeground(QColor("#8fd3ff"))
                    self.pair_table.setItem(row, 4, vr_item)
                # ── 实时刷新走势预判列 ──
                self._set_trend_cell(row, inst_id)
                break

    def refresh_pair_table(self):
        """刷新监控列表表格（置顶交易对排在最前）"""
        pairs = self.monitor_worker.get_pairs() if self.monitor_worker else []
        if not pairs:
            pairs = self.load_pairs_from_config()
        # 置顶排序：先在 _pinned_pairs 中的排前面
        pairs = sorted(pairs, key=lambda p: (0 if p in self._pinned_pairs else 1, p))
        # 保存当前列宽，避免 setRowCount(0) 在 macOS 上触发布局重置
        saved_widths = [self.pair_table.columnWidth(i) for i in range(self.pair_table.columnCount())]
        self.pair_table.setRowCount(0)
        for inst_id in pairs:
            row = self.pair_table.rowCount()
            self.pair_table.insertRow(row)
            self.pair_table.setItem(row, 0, QTableWidgetItem(inst_id))
            if inst_id in self._pinned_pairs:
                # 置顶行背景高亮
                pin_item = self.pair_table.item(row, 0)
                if pin_item:
                    pin_item.setForeground(QColor("#ffd700"))
            self.pair_table.setItem(row, 1, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 2, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 3, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 4, QTableWidgetItem("--"))
            self.pair_table.setItem(row, 5, QTableWidgetItem("待启动"))
            self.pair_table.setItem(row, 6, QTableWidgetItem("--"))
            self._set_trend_cell(row, inst_id)
        self.pair_count_label.setText(f"{len(pairs)} 个交易对")
        # 恢复列宽（setRowCount(0) 可能在部分平台上触发布局重置）
        if saved_widths and len(saved_widths) == self.pair_table.columnCount():
            for i, w in enumerate(saved_widths):
                if w >= 50:
                    self.pair_table.setColumnWidth(i, int(w))

    # ── 定时推送监控报表 ──
    def _load_push_config(self):
        """加载推送配置"""
        try:
            if self._push_config_path.exists():
                with open(self._push_config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                self.push_interval_spin.setValue(cfg.get('interval_min', 30))
                self.push_token_edit.setText(cfg.get('bot_token', ''))
                self.push_chat_id_edit.setText(cfg.get('chat_id', ''))
                if cfg.get('push_enabled', False):
                    QTimer.singleShot(1000, self.start_monitor_push)
        except Exception:
            pass

    def _save_push_config(self):
        """保存推送配置"""
        try:
            cfg = {
                'interval_min': self.push_interval_spin.value(),
                'bot_token': self.push_token_edit.text().strip(),
                'chat_id': self.push_chat_id_edit.text().strip(),
                'push_enabled': self._push_timer.isActive(),
            }
            with open(self._push_config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def start_monitor_push(self):
        """启动定时推送"""
        interval = max(1, self.push_interval_spin.value())
        if not self.push_token_edit.text().strip() or not self.push_chat_id_edit.text().strip():
            QMessageBox.warning(self, "配置缺失", "请填写 Telegram Bot Token 和 Chat ID")
            return
        self._push_timer.start(interval * 60 * 1000)
        self.push_start_btn.setEnabled(False)
        self.push_stop_btn.setEnabled(True)
        self.push_status_label.setText(f"🟢 每 {interval} 分钟推送")
        self.push_status_label.setStyleSheet("color: #00ff88; font-size: 11px;")
        self._save_push_config()

    def stop_monitor_push(self):
        """停止定时推送"""
        self._push_timer.stop()
        self.push_start_btn.setEnabled(True)
        self.push_stop_btn.setEnabled(False)
        self.push_status_label.setText("⏸ 已停止")
        self.push_status_label.setStyleSheet("color: #aaaaaa; font-size: 11px;")
        self._save_push_config()

    def _do_monitor_push(self):
        """执行一次监控报表推送（线程安全：数据在主线程收集）"""
        # 在主线程收集表格数据后传递给后台线程
        rows_data = []
        for row in range(self.pair_table.rowCount()):
            row_data = []
            for col in range(8):
                item = self.pair_table.item(row, col)
                row_data.append(item.text() if item else "--")
            rows_data.append(row_data)
        import threading
        threading.Thread(target=lambda: self._monitor_push_worker(rows_data), daemon=True).start()

    def _monitor_push_worker(self, rows_data: list):
        """后台线程：生成 Excel 并发送到 Telegram（使用预收集的数据，不访问 Qt widgets）"""
        try:
            file_path = self._export_monitor_to_excel()
            if not file_path:
                return
            token = self.push_token_edit.text().strip()
            chat_id = self.push_chat_id_edit.text().strip()
            if not token or not chat_id:
                return
            caption = f"📊 重点监控池报表\n⏰ {datetime.now().strftime('%Y-%m-%d %H:%M')}\n📋 共 {self.pair_table.rowCount()} 个交易对"
            try:
                import requests
                url = f"https://api.telegram.org/bot{token}/sendDocument"
                with open(file_path, 'rb') as f:
                    resp = requests.post(url, data={'chat_id': chat_id, 'caption': caption}, files={'document': f}, timeout=30)
                if resp.status_code == 200 and resp.json().get('ok'):
                    QTimer.singleShot(0, lambda: self.on_log_message("监控报表已推送到 Telegram", "SUCCESS"))
                else:
                    QTimer.singleShot(0, lambda: self.on_log_message(f"推送失败: {resp.text[:100]}", "ERROR"))
            except Exception as e:
                QTimer.singleShot(0, lambda: self.on_log_message(f"推送异常: {e}", "ERROR"))
        except Exception as e:
            QTimer.singleShot(0, lambda: self.on_log_message(f"导出报表失败: {e}", "ERROR"))

    def _export_monitor_to_excel(self) -> Optional[str]:
        """导出监控池数据到 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            QTimer.singleShot(0, lambda: QMessageBox.warning(self, "缺少依赖", "请安装 openpyxl: pip install openpyxl"))
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MonitorPool"

        # 表头
        headers = ["交易对", "最新价", "涨跌24h%", "RSI 1H", "量比", "状态", "最近信号", "走势预判(日/时/分)"]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='333333'),
            right=Side(style='thin', color='333333'),
            top=Side(style='thin', color='333333'),
            bottom=Side(style='thin', color='333333'),
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 数据行（使用预收集的数据，不访问 Qt widgets）
        data_font = Font(color="CCCCCC", size=10)
        data_fill = PatternFill("solid", fgColor="1A1A1A")
        for ri, row_data in enumerate(rows_data):
            for ci, val in enumerate(row_data):
                cell = ws.cell(row=ri + 2, column=ci + 1, value=val)
                cell.font = data_font
                cell.fill = data_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # 列宽
        widths = [14, 12, 10, 8, 8, 10, 14, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        export_dir = self._config_dir / "reports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = f"monitor_pool_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = str(export_dir / filename)
        wb.save(file_path)
        return file_path

    def closeEvent(self, event):
        """关闭时停止监控"""
        self._push_timer.stop()
        if self.monitor_worker and self.monitor_worker.isRunning():
            self.monitor_worker.stop()
            self.monitor_worker.wait(200)  # 非阻塞，让线程自然退出
        if self._signal_log_window:
            self._signal_log_window.close()
            self._signal_log_window = None
        super().closeEvent(event)
